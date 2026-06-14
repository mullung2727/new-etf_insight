"""Parse kiwoom_api.xlsx — dump a TR spec by code, or list all TRs.

Usage:
  python parse.py                 # list every TR: code -> Korean name
  python parse.py ka10075         # dump one TR field table
  python parse.py kt10002 kt10003 # dump several
  python parse.py --grep 미체결    # find TRs whose sheet name contains text

Windows note: forces UTF-8 stdout so Korean prints clean regardless of console codepage.
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

import openpyxl

# UTF-8 stdout (Windows consoles default to cp949 and mangle Korean).
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

# xlsx lives at the repo root; this file is .claude/skills/kiwoom-api/parse.py
XLSX = Path(__file__).resolve().parents[3] / "kiwoom_api.xlsx"


def _wb():
    if not XLSX.exists():
        sys.exit(f"not found: {XLSX}")
    return openpyxl.load_workbook(XLSX, read_only=True)


def list_all(wb) -> None:
    for name in wb.sheetnames:
        print(name)


def grep(wb, term: str) -> None:
    for name in wb.sheetnames:
        if term in name:
            print(name)


def dump(wb, code: str) -> None:
    matches = [n for n in wb.sheetnames if code in n]
    if not matches:
        print(f"!! no sheet for {code}")
        return
    ws = wb[matches[0]]
    print("#" * 12, matches[0])
    for row in ws.iter_rows(values_only=True):
        cells = [str(c) for c in row if c is not None and str(c).strip()]
        if cells:
            print(" | ".join(cells))
    print()


def main() -> None:
    args = sys.argv[1:]
    wb = _wb()
    if not args:
        list_all(wb)
    elif args[0] == "--grep":
        grep(wb, args[1])
    else:
        for code in args:
            dump(wb, code)


if __name__ == "__main__":
    main()
